import openpyxl

import config


def extract_data_from_file(file_path):
    """
    Reads the Excel file and extracts a dictionary mapping 'name' to 'price'.
    Uses openpyxl in read-only mode for maximum performance and lower memory usage.
    """
    try:
        # Load workbook in read-only and data-only mode (ignores formulas, gets values)
        wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
        ws = wb.active

        # Extract headers from the first row
        headers = [str(cell.value).lower().strip() if cell.value is not None else "" for cell in ws[1]]

        try:
            name_idx = headers.index(config.name_column_header)
            price_idx = headers.index(config.price_column_header)
        except ValueError:
            # One or both required columns are missing
            wb.close()
            return {}

        extracted_data = {}

        # Iterate through data rows (skipping the header row)
        for row in ws.iter_rows(min_row=2, values_only=True):
            # Ensure the row has enough columns
            if len(row) > max(name_idx, price_idx):
                name_val = row[name_idx]
                price_val = row[price_idx]

                # Drop rows where either name or price is missing
                if name_val is not None and price_val is not None:
                    extracted_data[str(name_val)] = price_val

        wb.close()
        return extracted_data

    except Exception as e:
        print(f'Failed to read {file_path}: {e}')
        return {}


def run_pipeline(totals):
    """
    Receives a list of calculated totals from the UI.
    Returns a formatted string containing the grand total.
    """
    grand_total = sum(totals)

    results = [
        # f'Summing {len(totals)} rows...\n',
        # '-' * 30 + '\n',
        f'Total: {grand_total:.2f}\n',
        '-' * 30 + '\n',
        '✅ Processing complete!'
    ]

    return ''.join(results)
